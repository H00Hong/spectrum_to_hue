# -*- coding: utf-8 -*-
import csv
import os.path
from typing import Dict, List, Union

import about
import openpyxl
import pdfplumber
import wx
import xlrd
from _base import (WIDGETS_TOTAL, line, line_h, line_v, get_setting,
                   read_pdf, read_txt, save_setting)
from cie import CIE
from mywxwidgets.grid.gridnumpy import Grid, GridWithHeader
from numpy import array, c_, ndarray, vstack


class CIE(CIE):

    def colour(self):
        XYZ = self.spec2xyz()
        if XYZ.ndim == 1:
            XYZ = XYZ[:, None]
        lab = self.xyz2lab(XYZ)
        hlab = self.xyz2lab_h(XYZ)
        rgb = self.xyz2rgb(XYZ)
        uv_ = self.xyz2yuv(XYZ)[1:]
        luv = self.xyz2luv(XYZ)
        # rgb16_ = self.rgb16(rgb, 1)
        xyz = XYZ / XYZ.sum(axis=0)
        # uvw_ = np.asarray([*uv_, 1 - uv_.sum(axis=0)])
        Ch_ab = self.chs(lab)[0:2]
        Ch_ab_h = self.chs(hlab)[0:2]
        Chs_uv = self.chs(luv)
        yi = 100*(1.28*XYZ[0]-1.06*XYZ[2])/XYZ[1]
        return vstack((XYZ, xyz, lab, Ch_ab, hlab, Ch_ab_h, self.rgb16(rgb, 1),
                       uv_, 1 - uv_.sum(axis=0), luv, Chs_uv, yi))


CHECKBOX_TITLE = ['X', 'Y', 'Z', 'x', 'y', 'z', 'CIELAB-L*', 'CIELAB-a*',
                  'CIELAB-b*', 'CIELAB-C*_ab', 'CIELAB-h_ab', 'Hunter L',
                  'Hunter a', 'Hunter b', 'Hunter C_ab', 'Hunter h_ab', 'sRGB',
                  'u\'', 'v\'', 'w\'', 'CIELUV-L*', 'CIELUV-u*', 'CIELUV-v*',
                  'CIELUV-C*_uv', 'CIELUV-h_uv', 'CIELUV-s_uv']
CHECKBOX_LABEL = ['X', 'Y', 'Z', 'x', 'y', 'z', 'L*', 'a*', 'b*', 'C*_ab',
                 'h_ab', 'L', 'a', 'b', 'C_ab', 'h_ab', 'sRGB', 'u\'', 'v\'',
                 'w\'', 'L*', 'u*', 'v*', 'C*_uv', 'h_uv', 's_uv']
# CHECKBOX_BOOL = (True, True, True, False, False, False, True, True, True,
#                  False, False, False, False, False, False, False, True, False,
#                  False, False, False, False, False, False, False, False)
CHECKBOX_DICT = dict(zip(CHECKBOX_TITLE, CHECKBOX_LABEL))


class CalcItems(wx.Dialog):

    def __init__(self, parent, c: bool) -> None:
        super(CalcItems, self).__init__(parent,
                                        title='选择计算项目',
                                        size=(600, 600))
        self.setting = get_setting()
        self._c = c
        self._init_ui()
        self.Centre()

    def _init_ui(self):
        self.lab3 = wx.StaticText(self, label='CIE XYZ')
        self.lab6 = wx.StaticText(self, label='CIE xyz')
        self.lab9 = wx.StaticText(self, label='CIELAB')
        self.lab12 = wx.StaticText(self, label='Hunter Lab')
        self.lab17 = wx.StaticText(self, label='CIE u\'v\'w\'')
        self.lab20 = wx.StaticText(self, label='CIELUV')
        self.btn231 = wx.Button(self, label='确定')
        self.btn232 = wx.Button(self, label='取消')

        self.check_boxs = {
            k: wx.CheckBox(self, label=l)
            for k, l in CHECKBOX_DICT.items()
        }
        if self._c:
            self.cb_yi = wx.CheckBox(self, label='YI')
            self.cb_yi.SetValue(self.setting['YI'])
        for k, v in self.check_boxs.items():
            v.SetValue(self.setting[k])
        self._set_font()
        self._laypout_0()

        self.Bind(wx.EVT_BUTTON, self._on_ok, self.btn231)
        self.Bind(wx.EVT_BUTTON, self._on_cancel, self.btn232)

    def _on_ok(self, event):
        self.result: Dict[str, bool] = {}
        for k, v in self.check_boxs.items():
            b: bool = v.GetValue()
            self.setting[k] = b
            self.result[k] = b

        b = self.cb_yi.GetValue() if self._c else False
        self.setting['YI'] = b
        self.result['YI'] = b
        save_setting(self.setting)
        self.EndModal(wx.ID_OK)  # 这里传入的参数会在 ShowModal return

    def _on_cancel(self, event):
        self.EndModal(wx.ID_CANCEL)  # 这里传入的参数会在 ShowModal return

    def _layout_1(self, layout1: wx.BoxSizer):
        layout10 = wx.BoxSizer(wx.VERTICAL)
        layout10.Add(line_h(self), 0, wx.EXPAND | wx.ALL, 3)
        layout11 = wx.BoxSizer(wx.VERTICAL)
        layout11.Add(line_h(self), 0, wx.EXPAND | wx.ALL, 3)

        for lab, cb_t in zip(
            (self.lab3, self.lab6, self.lab17),
            (('X', 'Y', 'Z'), ('x', 'y', 'z'), ('u\'', 'v\'', 'w\''))):
            layout10.Add(lab, 0, wx.ALL, 3)
            layout = wx.BoxSizer(wx.HORIZONTAL)
            for i in cb_t:
                layout.Add(self.check_boxs[i], 1, wx.ALL | wx.EXPAND, 3)
            layout10.Add(layout, 1, wx.EXPAND | wx.ALL, 3)
            layout10.Add(line_h(self), 0, wx.EXPAND | wx.ALL, 3)

        for lab, cb_t in zip(
            (self.lab9, self.lab12, self.lab20),
            (('CIELAB-L*', 'CIELAB-a*', 'CIELAB-b*', 'CIELAB-C*_ab',
              'CIELAB-h_ab'), ('Hunter L', 'Hunter a', 'Hunter b',
                               'Hunter C_ab', 'Hunter h_ab'),
             ('CIELUV-L*', 'CIELUV-u*', 'CIELUV-v*', 'CIELUV-C*_uv',
              'CIELUV-h_uv', 'CIELUV-s_uv'))):
            layout11.Add(lab, 0, wx.ALL, 3)
            layout = wx.GridSizer(0, 3, 0, 0)
            for i in cb_t:
                layout.Add(self.check_boxs[i], 1, wx.ALL, 3)
            layout11.Add(layout, 1, wx.EXPAND | wx.ALL, 3)
            layout11.Add(line_h(self), 0, wx.EXPAND | wx.ALL, 3)

        layout1.Add(layout10, 1, wx.EXPAND | wx.ALL, 3)
        layout1.Add(line_v(self), 0, wx.EXPAND | wx.ALL, 3)
        layout1.Add(layout11, 1, wx.EXPAND | wx.ALL, 3)

    def _laypout_0(self):
        layout0 = wx.BoxSizer(wx.VERTICAL)
        layout1 = wx.BoxSizer(wx.HORIZONTAL)
        self._layout_1(layout1)
        layout0.Add(layout1, 1, wx.EXPAND | wx.ALL, 3)

        if self._c:
            layout = wx.BoxSizer(wx.HORIZONTAL)
            layout.Add(self.check_boxs['sRGB'], 0, wx.ALL, 3)
            layout.Add(self.cb_yi, 0, wx.ALL, 3)
            layout0.Add(layout, 0, wx.ALL)
        else:
            layout0.Add(self.check_boxs['sRGB'], 0, wx.ALL, 3)
        layout0.Add(line_h(self), 0, wx.EXPAND | wx.ALL, 3)

        layout1 = wx.BoxSizer(wx.HORIZONTAL)
        layout1.Add(wx.StaticText(self), 1, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 3)
        layout1.Add(self.btn231, 0, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 3)
        layout1.Add(self.btn232, 0, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 3)
        layout0.Add(layout1, 0, wx.EXPAND | wx.ALL, 3)
        self.SetSizer(layout0)

    def _set_font(self):
        FONT0 = wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL,
                        wx.FONTWEIGHT_NORMAL, False, 'Microsoft Yahei')
        self.SetFont(FONT0)
        # self.lab0.SetFont(FONT0)
        self.lab3.SetFont(FONT0)
        self.lab6.SetFont(FONT0)
        self.lab9.SetFont(FONT0)
        self.lab12.SetFont(FONT0)
        self.lab17.SetFont(FONT0)
        self.lab20.SetFont(FONT0)
        for v in self.check_boxs.values():
            v.SetFont(FONT0)
        self.btn231.SetFont(FONT0)
        self.btn232.SetFont(FONT0)
        if self._c:
            self.cb_yi.SetFont(FONT0)


class ReadFileData(wx.Dialog):

    def __init__(self, parent, path: str, grid: GridWithHeader):
        super(ReadFileData, self).__init__(parent, title='确认导入参数')
        self._path, self._grid = path, grid
        self.setting = get_setting()
        self.FONT0 = wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL,
                             wx.FONTWEIGHT_NORMAL, False, 'Microsoft Yahei')
        self._suffix = self._path.rpartition('.')[-1].lower()
        self._init_ui()
        if self._suffix == 'pdf':
            self._init_ui_0()
        else:
            self._init_ui_1()

    def _add_line_h(self):
        self.layout0.Add(line_h(self), 0, wx.ALL | wx.EXPAND, 3)

    def _init_ui(self):
        self.lab0 = wx.StaticText(self, label='请确认')
        self.lab1 = wx.StaticText(self, label='导入文件:')
        self.lab2 = wx.StaticText(self, label=self._path)
        self.lab3_00 = wx.StaticText(self, label='')
        self.sc3_10 = wx.SpinCtrl(self)

        self.lab_none = wx.StaticText(self, label='')
        self.btn_ok = wx.Button(self, label='确认')
        self.btn_cancel = wx.Button(self, label='取消')

        self.lab0.SetFont(self.FONT0)
        self.lab1.SetFont(self.FONT0)
        self.lab2.SetFont(self.FONT0)
        self.lab3_00.SetFont(self.FONT0)
        self.sc3_10.SetFont(self.FONT0)
        self.btn_ok.SetFont(self.FONT0)
        self.btn_cancel.SetFont(self.FONT0)

        self.layout0 = wx.BoxSizer(wx.VERTICAL)
        self.layout0.Add(self.lab0, 0, wx.ALL, 3)
        self._add_line_h()
        self.layout0.Add(self.lab1, 0, wx.ALL, 3)
        self.layout0.Add(self.lab2, 1, wx.ALL | wx.EXPAND, 3)
        self._add_line_h()

        self.layout1 = wx.BoxSizer(wx.HORIZONTAL)
        self.layout1.Add(self.lab_none, 1, wx.ALL | wx.EXPAND, 3)
        self.layout1.Add(self.btn_ok, 0, wx.ALL | wx.EXPAND, 3)
        self.layout1.Add(self.btn_cancel, 0, wx.ALL | wx.EXPAND, 3)

        self.Bind(wx.EVT_BUTTON, self._on_ok, self.btn_ok)
        self.Bind(wx.EVT_BUTTON, self._on_cancel, self.btn_cancel)

    def _init_ui_1(self):
        self.lab3_00.SetLabel('请确认 表头/名称在第几行\n  0表示没有')
        self.lab3_10 = wx.StaticText(self, label='确认导入数据在哪个数据表')
        self.lab3_20 = wx.StaticText(self, label='确认导入数据波长在第几列')

        if self._suffix in ('csv', 'txt'):
            xl1 = [self._path.rpartition('/')[-1][:-4]] # 读取文件名
        elif self._suffix == 'xlsx':
            # 读取Excel的sheetname
            wb = openpyxl.load_workbook(self._path)
            xl1 = wb.sheetnames
        else:
            wb = xlrd.open_workbook(self._path)
            xl1 = wb.sheet_names()

        self.cb3_11 = wx.Choice(self, choices=xl1)
        self.cb3_11.SetSelection(0)
        self.sc3_12 = wx.SpinCtrl(self, min=1)

        self.sc3_10.SetValue(self.setting['header_row'])
        self.sc3_12.SetValue(self.setting['wavelength_col'])

        self.lab3_10.SetFont(self.FONT0)
        self.lab3_20.SetFont(self.FONT0)
        self.cb3_11.SetFont(self.FONT0)
        self.sc3_12.SetFont(self.FONT0)

        layout2 = wx.GridSizer(0, 2, 0, 0)
        layout2.Add(self.lab3_00, 1, wx.EXPAND | wx.ALIGN_CENTER_VERTICAL, 3)
        layout2.Add(self.sc3_10, 1, wx.EXPAND | wx.ALIGN_CENTER_VERTICAL, 3)
        layout2.Add(self.lab3_10, 1, wx.EXPAND | wx.ALIGN_CENTER_VERTICAL, 3)
        layout2.Add(self.cb3_11, 1, wx.EXPAND | wx.ALIGN_CENTER_VERTICAL, 3)
        layout2.Add(self.lab3_20, 1, wx.EXPAND | wx.ALIGN_CENTER_VERTICAL, 3)
        layout2.Add(self.sc3_12, 1, wx.EXPAND | wx.ALIGN_CENTER_VERTICAL, 3)

        self.layout0.Add(layout2, 0, wx.ALL | wx.EXPAND, 3)
        self._add_line_h()
        self.layout0.Add(self.layout1, 0, wx.ALL | wx.EXPAND, 3)
        self.SetSizer(self.layout0)

        self.SetSize((500, 410))

    def _init_ui_0(self):
        self.lab3_00.SetLabel('确认数据横跨了几页')
        self.sc3_10.SetMax(len(pdfplumber.open(self._path).pages) - 1)
        self.sc3_10.SetValue(self.setting['pdf_page'])

        layout2 = wx.BoxSizer(wx.HORIZONTAL)
        layout2.Add(self.lab3_00, 1, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 3)
        layout2.Add(self.sc3_10, 1, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 3)
        self.layout0.Add(layout2, 0, wx.ALL | wx.EXPAND, 3)
        self._add_line_h()
        self.layout0.Add(self.layout1, 0, wx.ALL | wx.EXPAND, 3)
        self.SetSizer(self.layout0)

        self.SetSize((300, 310))

    def _on_ok(self, event):
        if self._suffix == 'pdf':
            self.setting['pdf_page'] = self.sc3_10.GetValue()
        else:
            self.setting['header_row'] = self.sc3_10.GetValue()
            self.setting['wavelength_col'] = self.sc3_12.GetValue()
        save_setting(self.setting)
        self._read_data()
        wx.LogMessage('导入完成')
        self.EndModal(wx.ID_OK)

    def _read_data(self):
        ind1 = int(self.sc3_10.GetValue())  # 读取pdf的几页 or 读取Excel的第几行
        if self._suffix == 'pdf':
            spe = read_pdf(self._path, ind1)
            hea = ['波长'] + [f'Data{i}' for i in range(1, len(spe[0]))]
        else:
            ind2 = self.cb3_11.GetStringSelection()  # 读取Excel的sheetname
            ind3 = int(self.sc3_12.GetValue()) - 1  # 读取Excel的第几列
            ind1_0 = None if ind1 == 0 else ind1 - 1
            if self._suffix in ('csv', 'txt', 'tsv'):
                spe, hea = read_txt(self._path, ind1_0, slice(ind3, None))

            elif self._suffix == 'xlsx':
                wb1 = openpyxl.load_workbook(self._path)
                ws1 = wb1[ind2]
                spe: List[List[str]] = list(
                    map(list,
                        ws1.iter_rows(min_row=ind1,
                                      min_col=ind3 + 1,
                                      values_only=True)))

                if ind1_0 is not None:
                    hea = spe.pop(0)

            else:
                wb2: xlrd.book.Book = xlrd.open_workbook(self._path)
                ws2 = wb2.sheet_by_name(ind2)
                if ind1_0 is not None:
                    hea = ws2.row_values(ind1_0, start_colx=ind3)
                spe = [
                    ws2.row_values(i, start_colx=ind3)
                    for i in range(ind1, ws2.nrows)
                ]

            if ind1_0 is None:
                hea = ['波长'] + [f'Data{i}' for i in range(1, len(spe[0]))]
            else:
                if isinstance(hea[0], list):
                    for i in range(len(hea)):
                        hea[i][0] = '波长'
                else:
                    hea[0] = '波长'

        self._grid.SetSubject(spe)
        self._grid.SetHeader([hea])

    def _on_cancel(self, event):
        wx.LogMessage('导入取消')
        self.EndModal(wx.ID_CANCEL)


WIDGETS_LABEL = WIDGETS_TOTAL['spec2hue']
MAP_L2W = {
    'label': wx.StaticText,
    'button': wx.Button,
    'text': wx.TextCtrl,
    'choice': wx.Choice,
    'GridWithHeader': GridWithHeader,
    'Grid': Grid,
    'line': line
}


class Spec2Hue(wx.Panel):
    widgets: Dict[str, Union[wx.StaticText, wx.Button, wx.TextCtrl, wx.Choice,
                             GridWithHeader, Grid, wx.StaticLine]]

    def __init__(self, parent):
        super(Spec2Hue, self).__init__(parent)
        self.SetBackgroundColour(wx.Colour(245, 245, 245))
        # self.SetSize((800, 500))
        # self.SetMinSize((680, 235))
        self._init_ui()
        self._set_bind()
        self.Centre()

    def _init_ui(self):
        self.widgets = {}

        def loop_append(dic: Dict[str, dict], out: dict):
            num = -1
            for k, v in dic.items():
                lab = k.split('_')
                lab0 = int(lab[0])
                if lab0 != num:
                    num = lab0
                    out.setdefault(lab0, list())
                widget = MAP_L2W[lab[1]](self, **v)
                out[lab0].append(widget)
                self.widgets['_'.join(lab[1:])] = widget

        widgets_left: Dict[int, list] = {}
        loop_append(WIDGETS_LABEL['left'], widgets_left)
        widgets_right: Dict[int, list] = {}
        loop_append(WIDGETS_LABEL['right'], widgets_right)

        self.grid_in: GridWithHeader = self.widgets['GridWithHeader_spectrum']
        self.grid_in.SetHeader([['波长', '', '']])
        self.grid_in.SetHeaderLabels([f'标题{i+1}' for i in range(25)])
        self.filepath: wx.TextCtrl = self.widgets['text_import']
        self.widgets['choice_wavelength.unit'].SetSelection(0)
        self.widgets['choice_spectrum.upper'].SetSelection(1)
        self.widgets['choice_si'].SetSelection(1)
        self.widgets['choice_vi'].SetSelection(0)
        self.grid_out: Grid = self.widgets['Grid_output']
        self.grid_out.HideRowLabels()
        self.grid_out.HideColLabels()

        self._set_font()
        layout = wx.BoxSizer(wx.HORIZONTAL)
        layout.Add(self._layout_left(widgets_left), 1, wx.EXPAND, 4)
        layout.Add(line_v(self), 0, wx.EXPAND | wx.ALL, 4)
        layout.Add(self._layout_right(widgets_right), 1, wx.EXPAND, 4)
        self.SetSizer(layout)

    def _layout_left(self, widgets_left: Dict[int, list]):
        layout_left = wx.BoxSizer(wx.VERTICAL)
        for i in range(3):
            widgets = widgets_left[i]
            layout = wx.BoxSizer(wx.HORIZONTAL)
            for ii, w in enumerate(widgets):
                layout.Add(w, int(ii == 1), wx.ALIGN_CENTER_VERTICAL | wx.ALL, 4)
            layout_left.Add(layout, 0, wx.EXPAND, 4)
        layout_left.Add(self.grid_in, 1, wx.ALL | wx.EXPAND, 4)
        layout = wx.BoxSizer(wx.HORIZONTAL)
        layout.Add(widgets_left[4][0], 1, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 4)
        for w in widgets_left[4][1:]:
            layout.Add(w, 0, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 4)
        layout_left.Add(layout, 0, wx.EXPAND, 4)
        return layout_left

    def _layout_right(self, widgets_right: Dict[int, list]):
        layout_right = wx.BoxSizer(wx.VERTICAL)
        layout_right.Add(widgets_right[0][0], 0, wx.ALL, 4)
        layout = wx.BoxSizer(wx.HORIZONTAL)
        for i in range(3):
            layout.Add(widgets_right[1][i], 0, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 4)
        layout.Add(widgets_right[1][3], 0, wx.EXPAND, 4)
        layout.Add(widgets_right[1][4], 1, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 4)
        layout.Add(widgets_right[1][5], 0, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 4)
        layout_right.Add(layout, 0, wx.EXPAND, 4)
        layout022 = wx.BoxSizer(wx.HORIZONTAL)
        layout022.Add(widgets_right[2][0], 0, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 4)
        layout022.Add(widgets_right[2][1], 1, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 4)
        layout_right.Add(layout022, 0, wx.EXPAND, 4)
        layout_right.Add(self.grid_out, 1, wx.ALL | wx.EXPAND, 4)
        layout023 = wx.BoxSizer(wx.HORIZONTAL)
        layout023.Add(widgets_right[4][0], 1, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 4)
        layout023.Add(widgets_right[4][1], 0, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 4)
        layout_right.Add(layout023, 0, wx.EXPAND, 4)
        return layout_right

    def _set_font(self):
        FONT0 = wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL,
                        wx.FONTWEIGHT_NORMAL, False, 'Microsoft Yahei')
        FONT1 = wx.Font(16, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL,
                        wx.FONTWEIGHT_NORMAL, False, 'Microsoft Yahei')
        FONT2 = wx.Font(13, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL,
                        wx.FONTWEIGHT_NORMAL, False, 'Microsoft Yahei')

        for v in self.widgets.values():
            v.SetFont(FONT0)
        for i in ['label_input', 'label_spectrum', 'label_output', 'label_calculate']:
            self.widgets[i].SetFont(FONT1)
        self.widgets['button_instructions'].SetFont(FONT2)

    def _set_bind(self):
        self.widgets['button_instructions'].Bind(wx.EVT_BUTTON, self._on_btn_instructions)  # 操作说明
        self.widgets['button_select'].Bind(wx.EVT_BUTTON, self._on_btn_filepath)  # 选择文件
        self.widgets['button_import'].Bind(wx.EVT_BUTTON, self._on_btn_import)  # 导入
        self.widgets['button_calculate'].Bind(wx.EVT_BUTTON, self._on_btn_calc)  # 计算
        self.widgets['button_save'].Bind(wx.EVT_BUTTON, self._on_btn_save)  # 导出
        self.Bind(wx.EVT_SIZE, self._on_size)

    def _on_size(self, event):
        # 配合改变高度
        w, _ = event.GetSize()
        w = w // 2 - 5
        self.grid_in.SetMinSize((w, -1))
        self.grid_in.SetMaxSize((w, -1))
        event.Skip()

    def _on_btn_instructions(self, event):
        # 操作说明
        win = about.AboutWin(self)
        win.Show()

    def _on_btn_filepath(self, event):
        # 选择文件
        try:
            setting = get_setting()
            wildcard = 'Data file (*.xlsx;*.xls;*.csv;*.txt;*.tsv;*.pdf)|*.xlsx;*.xls;*.csv;*.txt;*.tsv;*.pdf|\
Excel files (*.xlsx;*.xls)|*.xlsx;*.xls|\
CSV files (*.csv;*.txt;*.tsv)|*.csv;*.txt;*.tsv|PDF file (*.pdf)|*.pdf'
            fileDialog = wx.FileDialog(self,
                                       '选择数据文件',
                                       defaultDir=setting['open_path'],
                                       wildcard=wildcard,
                                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST)

            if fileDialog.ShowModal() != wx.ID_OK:
                return
            path: str = fileDialog.GetPath().replace('\\', '/')
            self.filepath.SetValue(path)

            setting['open_path'] = os.path.dirname(path)
            save_setting(setting)
        except Exception as e:
            self._err(e)

    def _on_btn_import(self, event):
        # 导入
        try:
            path: str = self.filepath.GetValue().replace('\\', '/')
            readfile = ReadFileData(self, path, self.grid_in)
            if readfile.ShowModal() != wx.ID_OK:
                return
            self.Layout()

            setting = get_setting()
            setting['open_path'] = os.path.dirname(path)
            save_setting(setting)
        except Exception as e:
            self._err(e)

    def _on_btn_calc(self, event):
        # 计算
        try:
            header = array(self.grid_in.header.dataBase.data, str)
        except Exception as e:
            self._err(f'header 格式不正确:\n{e}')
            return
        for i in range(len(header)):
            header[i, 0] = '-'
        try:
            spe = array(self.grid_in.subject.dataBase.data, float)
        except Exception as e:
            if str(e).endswith("('')"):
                self._err('数据中不能有空单元格, 如果必须请用 0 代替')
            else:
                self._err(f'数据格式不正确, 必须全部是数字:\n{e}')
            return

        si = self.widgets['choice_si'].GetStringSelection()
        choose = CalcItems(self, si=='C')
        if choose.ShowModal() != wx.ID_OK:
            return

        try:
            hue_hea_b = choose.result
            hue = CIE(spe, si,
                      self.widgets['choice_vi'].GetSelection() * 8 + 2,
                      self.widgets['choice_wavelength.unit'].GetStringSelection(),
                      int(self.widgets['choice_spectrum.upper'].GetStringSelection())
                      ).colour()
            # hea_ = [hue_hea[i] for i, v in enumerate(hue_hea_b) if v]
            hue_ = c_[CHECKBOX_LABEL+['YI'], hue][list(hue_hea_b.values())]
            self.hue: ndarray[str] = vstack([header, hue_])

            self.grid_out.SetShowFormat('{:.3f}')
            self.grid_out.SetData(self.hue)
            self.grid_out.AutoSizeColumns()
            nc = len(self.hue[0])
        except Exception as e:
            self._err(e)
            return

        for ir in range(len(self.hue)):
            for ic in range(nc):
                # 初始化-覆盖上一次的颜色设置
                self.grid_out.SetCellBackgroundColour(ir, ic, wx.WHITE)
                self.grid_out.SetCellTextColour(ir, ic, wx.BLACK)
        hea_ = hue_[:, 0].tolist()
        if 'sRGB' in hea_:
            nr = hea_.index('sRGB') + 1
            for i in range(1, nc):
                val: str = self.hue[nr, i]
                r, g, b = int(val[1:3], 16), int(val[3:5], 16), int(val[5:7], 16)
                self.grid_out.SetCellBackgroundColour(nr, i, wx.Colour(r, g, b))
                gray = 0.299 * r + 0.578 * g + 0.114 * b
                c = wx.WHITE if gray < 0.78 else wx.BLACK
                self.grid_out.SetCellTextColour(nr, i, c)

    def _on_btn_save(self, event):
        # 导出
        path = os.path.dirname(self.filepath.GetValue())
        fileDialog = wx.FileDialog(self,
                                   '选择数据文件',
                                   defaultDir=path,
                                   wildcard='CSV files (*.csv)|*.csv',
                                   style=wx.FD_OPEN)
        if fileDialog.ShowModal() != wx.ID_OK:
            return
        path: str = fileDialog.GetPath()
        # print(path)
        try:
            with open(path, 'w', encoding='utf-8', newline='') as f:
                csv_writer = csv.writer(f)
                csv_writer.writerows(self.hue)
            wx.MessageBox('保存完毕', '提示', wx.OK | wx.ICON_INFORMATION)
        except Exception as e:
            self._err(e)

    def _err(self, text):
        wx.MessageBox(text, '警告', wx.OK | wx.ICON_ERROR)


if __name__ == '__main__':
    app = wx.App()
    win = wx.Frame(None, title='w2c', size=(800, 500))
    main = Spec2Hue(win)
    win.SetMinSize((680, 235))
    win.Show()
    app.MainLoop()
