# -*- coding: utf-8 -*-
import csv
import json
import os.path
from itertools import chain
from typing import Dict, List, Literal, Tuple, Union

import chardet
import openpyxl
import pdfplumber
import wx
import xlrd


def get_encoding(path: str) -> str:
    """读取csv的编码"""
    detect = chardet.detect(open(path, 'rb').read(4096))
    return detect['encoding']


def get_delimiter(path: str) -> Tuple[str, str]:
    """读取csv的分隔符 return (delimiter, encoding)"""
    sniffer = csv.Sniffer()
    encoding = get_encoding(path)
    with open(path, encoding=encoding) as fp:
        delimiter = sniffer.sniff(fp.read(4096)).delimiter
    return delimiter, encoding


def read_txt(
    path: str,
    header: Union[int, None] = None,
    col: slice = slice(0, None, None)
) -> Tuple[List[List[str]], List[str]]:
    """
    读取csv txt

    Parameters
    ----------
    path: str
        地址
    header: int | None
        表头行号 int 从0开始计数; 若无表头None
    col: slice
        列读取范围

    Returns
    -------
    data: List[List[str]]
        读取的数据
    header: List[str]
        表头
    """
    delimiter, encoding = get_delimiter(path)
    with open(path, encoding=encoding, newline='') as csvfile:
        csv_reader = csv.reader(csvfile, delimiter=delimiter)
        if header is None:
            data_header = []
        elif isinstance(header, int):
            for _ in range(header):
                next(csv_reader)
            data_header = next(csv_reader)[col]
        else:
            raise ValueError('header')

        data = [row[col] for row in csv_reader]
    return data, data_header


def read_pdf(path: str, data_ye: int = 0) -> List[List[str]]:
    """
    读取表格pdf
    data_ye为数据跨了几页 0为数据单页
    """
    assert isinstance(data_ye, int)
    data_ye += 1
    arr_t, arr_0, num = [], [], 0
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            del table[0]
            arr_0.append(table)
            num += 1
            if num % data_ye == 0:
                arr_t.append(list(chain.from_iterable(arr_0)))
                arr_0 = []
    pdf.flush_cache()  # 清理缓存 解除占用
    return list(chain.from_iterable(arr_t))


def line_h(parent):  # 横线
    return wx.StaticLine(parent, style=wx.LI_HORIZONTAL)


def line_v(parent):  # 竖线
    return wx.StaticLine(parent, style=wx.LI_VERTICAL)


def line(parent, style: Literal['h', 'v'] = 'h'):
    if style == 'h':
        return line_h(parent)
    elif style == 'v':
        return line_v(parent)


DIRNAME = os.path.dirname(__file__)
with open(DIRNAME + '/_widgets.json', 'r',
          encoding='utf-8') as fp:
    WIDGETS_TOTAL = json.load(fp)


def load_setting() -> Dict[str, Union[str, int, bool]]:
    with open(DIRNAME + '/_setting.json', 'r',
              encoding='utf-8') as fp:
        setting = json.load(fp)
    return setting


def save_setting(setting):
    with open(DIRNAME + '/_setting.json', 'w',
              encoding='utf-8') as fp:
        json.dump(setting, fp, ensure_ascii=False)


class ReadFileData(wx.Dialog):
    result = ()
    WIDGETS_LABEL = WIDGETS_TOTAL['readdata']

    def __init__(self, parent, path: str):
        super(ReadFileData, self).__init__(parent,
                                           style=wx.CAPTION | wx.CLOSE_BOX
                                           | wx.RESIZE_BORDER,
                                           title=self.WIDGETS_LABEL['title'])
        self._path = path
        self.setting = load_setting()
        self.widget_label = {}
        self.FONT0 = wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL,
                             wx.FONTWEIGHT_NORMAL, False, self.WIDGETS_LABEL['font'])
        self._suffix = self._path.rpartition('.')[-1].lower()
        self._init_ui()
        if self._suffix == 'pdf':
            self._init_ui_0()
        else:
            self._init_ui_1()

    def _add_line_h(self):
        self.layout0.Add(line_h(self), 0, wx.ALL | wx.EXPAND, 3)

    def _init_ui(self):
        self.lab_r0 = wx.StaticText(self, label=self.WIDGETS_LABEL['r0'])
        self.lab_r1 = wx.StaticText(self, label=self.WIDGETS_LABEL['r1'])
        self.lab_r2 = wx.StaticText(self, label=self._path)
        self.lab_r3_00 = wx.StaticText(self, label='')
        self.sc_r3_10 = wx.SpinCtrl(self)

        self.lab_none = wx.StaticText(self, label='')
        self.btn_ok = wx.Button(self, label='确定')
        self.btn_cancel = wx.Button(self, label='取消')

        self.lab_r0.SetFont(self.FONT0)
        self.lab_r1.SetFont(self.FONT0)
        self.lab_r2.SetFont(self.FONT0)
        self.lab_r3_00.SetFont(self.FONT0)
        self.sc_r3_10.SetFont(self.FONT0)
        self.btn_ok.SetFont(self.FONT0)
        self.btn_cancel.SetFont(self.FONT0)

        self.layout0 = wx.BoxSizer(wx.VERTICAL)
        self.layout0.Add(self.lab_r0, 0, wx.ALL, 3)
        self._add_line_h()
        self.layout0.Add(self.lab_r1, 0, wx.ALL, 3)
        self.layout0.Add(self.lab_r2, 1, wx.ALL | wx.EXPAND, 3)
        self._add_line_h()

        self.layout1 = wx.BoxSizer(wx.HORIZONTAL)
        self.layout1.Add(self.lab_none, 1, wx.ALL | wx.EXPAND, 3)
        self.layout1.Add(self.btn_ok, 0, wx.ALL | wx.EXPAND, 3)
        self.layout1.Add(self.btn_cancel, 0, wx.ALL | wx.EXPAND, 3)

        self.Bind(wx.EVT_BUTTON, self._on_ok, self.btn_ok)
        self.Bind(wx.EVT_BUTTON, self._on_cancel, self.btn_cancel)

    def _init_ui_1(self):
        self.lab_r3_00.SetLabel(self.WIDGETS_LABEL['r3_table0'])
        self.lab_r3_10 = wx.StaticText(self, label=self.WIDGETS_LABEL['r3_table1'])
        self.lab_r3_20 = wx.StaticText(self, label=self.WIDGETS_LABEL['r3_table2'])

        if self._suffix in ('csv', 'txt', 'tsv'):
            xl1 = [self._path.rpartition('/')[-1][:-4]] # 读取文件名
        elif self._suffix == 'xlsx':
            # 读取Excel的sheetname
            wb = openpyxl.load_workbook(self._path)
            xl1 = wb.sheetnames
        else:
            wb = xlrd.open_workbook(self._path)
            xl1 = wb.sheet_names()

        self.cb_r3_11 = wx.Choice(self, choices=xl1)
        self.cb_r3_11.SetSelection(0)
        self.sc_r3_12 = wx.SpinCtrl(self, min=1)

        self.sc_r3_10.SetValue(self.setting['header_row'])
        self.sc_r3_12.SetValue(self.setting['wavelength_col'])

        self.lab_r3_10.SetFont(self.FONT0)
        self.lab_r3_20.SetFont(self.FONT0)
        self.cb_r3_11.SetFont(self.FONT0)
        self.sc_r3_12.SetFont(self.FONT0)

        layout2 = wx.GridSizer(0, 2, 0, 0)
        layout2.Add(self.lab_r3_00, 1, wx.EXPAND | wx.ALIGN_CENTER_VERTICAL, 3)
        layout2.Add(self.sc_r3_10, 1, wx.EXPAND | wx.ALIGN_CENTER_VERTICAL, 3)
        layout2.Add(self.lab_r3_10, 1, wx.EXPAND | wx.ALIGN_CENTER_VERTICAL, 3)
        layout2.Add(self.cb_r3_11, 1, wx.EXPAND | wx.ALIGN_CENTER_VERTICAL, 3)
        layout2.Add(self.lab_r3_20, 1, wx.EXPAND | wx.ALIGN_CENTER_VERTICAL, 3)
        layout2.Add(self.sc_r3_12, 1, wx.EXPAND | wx.ALIGN_CENTER_VERTICAL, 3)

        self.layout0.Add(layout2, 0, wx.ALL | wx.EXPAND, 3)
        self._add_line_h()
        self.layout0.Add(self.layout1, 0, wx.ALL | wx.EXPAND, 3)
        self.SetSizer(self.layout0)
        self.SetSize((500, 410))

    def _init_ui_0(self):
        self.lab_r3_00.SetLabel(self.WIDGETS_LABEL['r3_pdf'])
        self.sc_r3_10.SetMax(len(pdfplumber.open(self._path).pages) - 1)
        self.sc_r3_10.SetValue(self.setting['pdf_page'])

        layout2 = wx.BoxSizer(wx.HORIZONTAL)
        layout2.Add(self.lab_r3_00, 1, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 3)
        layout2.Add(self.sc_r3_10, 1, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 3)
        self.layout0.Add(layout2, 0, wx.ALL | wx.EXPAND, 3)
        self._add_line_h()
        self.layout0.Add(self.layout1, 0, wx.ALL | wx.EXPAND, 3)
        self.SetSizer(self.layout0)

        self.SetSize((300, 310))

    def _on_ok(self, event):
        if self._suffix == 'pdf':
            self.setting['pdf_page'] = self.sc_r3_10.GetValue()
        else:
            self.setting['header_row'] = self.sc_r3_10.GetValue()
            self.setting['wavelength_col'] = self.sc_r3_12.GetValue()
        save_setting(self.setting)
        self._read_data()
        wx.LogMessage(self.WIDGETS_LABEL['message_ok'])
        self.EndModal(wx.ID_OK)

    def _read_data(self):
        ind1 = int(self.sc_r3_10.GetValue())  # 读取pdf的几页 or 读取Excel的第几行
        if self._suffix == 'pdf':
            spec = read_pdf(self._path, ind1)
            hea = ['波长'] + [f'Data{i}' for i in range(1, len(spec[0]))]
        else:
            ind2 = self.cb_r3_11.GetStringSelection()  # 读取Excel的sheetname
            ind3 = int(self.sc_r3_12.GetValue()) - 1  # 读取Excel的第几列
            ind1_0 = None if ind1 == 0 else ind1 - 1
            if self._suffix in ('csv', 'txt', 'tsv'):
                spec, hea = read_txt(self._path, ind1_0, slice(ind3, None))

            elif self._suffix == 'xlsx':
                wb1 = openpyxl.load_workbook(self._path)
                ws1 = wb1[ind2]
                spec: List[List[str]] = list(
                    map(list,
                        ws1.iter_rows(min_row=ind1,
                                      min_col=ind3 + 1,
                                      values_only=True)))

                if ind1_0 is not None:
                    hea = spec.pop(0)

            else:
                wb2: xlrd.book.Book = xlrd.open_workbook(self._path)
                ws2 = wb2.sheet_by_name(ind2)
                if ind1_0 is not None:
                    hea = ws2.row_values(ind1_0, start_colx=ind3)
                spec = [
                    ws2.row_values(i, start_colx=ind3)
                    for i in range(ind1, ws2.nrows)
                ]

            if ind1_0 is None:
                hea = ['波长'] + [f'Data{i}' for i in range(1, len(spec[0]))]
            else:
                if isinstance(hea[0], list):
                    for i in range(len(hea)):
                        hea[i][0] = '波长'
                else:
                    hea[0] = '波长'
        self.result = (spec, hea)
        # self._grid.SetSubject(spe)
        # self._grid.SetHeader([hea])

    def _on_cancel(self, event):
        wx.LogMessage(self.WIDGETS_LABEL['message_cancel'])
        self.EndModal(wx.ID_CANCEL)
